"""
DR Filing Autopilot — Streamlit App
Install:  pip install streamlit anthropic openpyxl pandas
Run:      streamlit run dr_filing_app.py
Secret:   ANTHROPIC_API_KEY = 'sk-ant-...'
"""

import io, json, re, os, time
from datetime import datetime
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import streamlit as st

st.set_page_config(page_title="DR Filing Autopilot", page_icon="📋", layout="wide")

# ── Excel headers — exactly matching template ─────────────────────────────────
STOCK_HEADERS = [
    "Run", "Company name", "Full company name", "Exchange name",
    "DR Ticker", "Units", "Ratio", "Address 1", "Address 2", "Tel", "Fax",
    "Company website", "Market name website", "Market website short",
    "UL Market website", "UL IR webiste", "UL IR News",
    "ATO fee", "", "Period",
    "Company name (Title Case)", "Latest Price",
]

ETF_HEADERS = [
    "Run", "ETF Name", "Exchange name", "Issuer",
    "DR Ticker", "Units", "Ratio", "Address 1", "Address 2", "Tel", "Fax",
    "ETF Website", "Market name website", "Market website ย่อ",
    "ETF Trading ticker", "Investment manager", "Distributor",
    "Underlying reference index", "Base currency", "Inception date",
    "Distribution frequency", "Distributor website",
    "ATO fee", "", "Period",
    "ETF Name (Title Case)", "Latest NAV/Price",
]

# ── System prompts ────────────────────────────────────────────────────────────
STOCK_SYSTEM_PROMPT = """You are a financial data assistant for a Thai DR (Depositary Receipt) issuer.
Given a stock ticker or company name, search the web for current data and return ONLY a valid JSON object — no markdown, no explanation.

Required keys:
{
  "companyName": "OFFICIAL NAME IN ALL CAPS e.g. TESLA INC.",
  "fullCompanyName": "ALL CAPS NAME + short nickname in double quotes e.g. TESLA INC. (\\"TESLA\\")",
  "companyNameTitle": "Title Case name e.g. Tesla Inc.",
  "latestPrice": "Current price with currency and date e.g. USD 248.50 (25 Feb 2026)",
  "exchangeName": "Thai exchange name + English — see mapping below",
  "exchangeNameEn": "English exchange name only e.g. NASDAQ",
  "drTicker": "Stock ticker + 80 e.g. TSLA80",
  "ratio": 1000,
  "address1": "Registered office address line 1",
  "address2": "City, State, Country, ZIP",
  "tel": "Phone with country code",
  "fax": "",
  "companyWebsite": "https://...",
  "marketNameWebsite": "Thai exchange name + English + (URL)",
  "marketWebsiteShort": "Exchange homepage URL",
  "ulMarketWebsite": "Direct stock quote page URL on the exchange",
  "ulIrWebsite": "Investor Relations page URL",
  "ulIrNews": "IR News / Press Releases URL",
  "atoFee": 0.4,
  "period": ""
}

ratio: integer — 100 if price<100, 1000 if price 100-999, 10000 if price>=1000
atoFee: 0.4 for all exchanges; 0.5 for Euronext Paris only

Thai exchange name mappings (use EXACTLY):
NYSE → นิวยอร์ก (NYSE) | https://www.nyse.com/
NASDAQ → แนสแด็ก (NASDAQ) | https://www.nasdaq.com/
Tokyo Stock Exchange → โตเกียว (Tokyo Stock Exchange) | https://www.jpx.co.jp/english/
HKEX → ฮ่องกง (The Stock Exchange of Hong Kong) เขตปกครองพิเศษฮ่องกง | https://www.hkex.com.hk/
Shanghai → เซี่ยงไฮ้ (Shanghai Stock Exchange) ประเทศจีน | https://english.sse.com.cn/home/
Shenzhen → เซิ้นเจิ้น (Shenzhen Stock Exchange) ประเทศจีน | https://www.szse.cn/English/index.html
Euronext Paris → ปารีส (Euronext Paris) | https://www.euronext.com/en/markets/paris
LSE → ลอนดอน (London Stock Exchange) | https://www.londonstockexchange.com/
NYSE Arca → นิวยอร์กอาร์ก้า (NYSE Arca) | https://www.nyse.com/markets/nyse-arca"""

ETF_SYSTEM_PROMPT = """You are a financial data assistant for a Thai DR (Depositary Receipt) issuer.
Given an ETF name or ticker, search the web for current data and return ONLY a valid JSON object — no markdown, no explanation.

Required keys:
{
  "etfName": "OFFICIAL ETF NAME IN ALL CAPS e.g. SPDR® GOLD TRUST",
  "etfNameTitle": "Title Case ETF name e.g. SPDR® Gold Trust",
  "exchangeName": "Thai exchange name + English — see mapping below",
  "exchangeNameEn": "English exchange name only",
  "issuer": "ETF issuer/trust name e.g. World Gold Trust Services LLC",
  "drTicker": "ETF ticker + 80 e.g. GOLDUS80",
  "ratio": 1000,
  "address1": "Issuer registered office address line 1",
  "address2": "City, State, Country, ZIP",
  "tel": "Phone with country code",
  "fax": "",
  "etfWebsite": "Official ETF or fund website URL",
  "marketNameWebsite": "Thai exchange name + English + URL on same/next line",
  "marketWebsiteShort": "Exchange homepage URL",
  "etfTradingTicker": "Exchange trading ticker e.g. GLD",
  "investmentManager": "Investment manager name",
  "distributor": "Distributor name",
  "underlyingIndex": "Underlying reference index or benchmark e.g. Price of gold bullion, less the Fund's expenses",
  "baseCurrency": "Base currency e.g. USD",
  "inceptionDate": "Fund inception date e.g. 18 Nov 2004",
  "distributionFrequency": "Distribution frequency e.g. Annually / None / Quarterly",
  "distributorWebsite": "Distributor website URL",
  "latestPrice": "Latest NAV or price with currency and date e.g. USD 248.50 (25 Feb 2026)",
  "atoFee": 0.4,
  "period": ""
}

ratio: integer — 100 if price<100, 1000 if price 100-999, 10000 if price>=1000
atoFee: 0.4 for all exchanges; 0.5 for Euronext Paris only

Thai exchange name mappings (use EXACTLY):
NYSE Arca → นิวยอร์กอาร์ก้า (NYSE Arca) | https://www.nyse.com/markets/nyse-arca
NYSE → นิวยอร์ก (NYSE) | https://www.nyse.com/
NASDAQ → แนสแด็ก (NASDAQ) | https://www.nasdaq.com/
Tokyo Stock Exchange → โตเกียว (Tokyo Stock Exchange) | https://www.jpx.co.jp/english/
HKEX → ฮ่องกง (The Stock Exchange of Hong Kong) เขตปกครองพิเศษฮ่องกง | https://www.hkex.com.hk/
LSE → ลอนดอน (London Stock Exchange) | https://www.londonstockexchange.com/"""


# ── API fetch ─────────────────────────────────────────────────────────────────
def fetch_data(query: str, api_key: str, system_prompt: str) -> dict:
    client = anthropic.Anthropic(api_key=api_key)
    for attempt in range(4):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1000,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                system=system_prompt,
                messages=[{"role": "user", "content": f"DR filing data for: {query}"}],
            )
            text = next((b.text for b in response.content if b.type == "text"), "")
            text = re.sub(r"```json\s*|```\s*", "", text).strip()
            m = re.search(r'\{.*\}', text, re.DOTALL)
            if not m:
                raise ValueError("No JSON in response — please try again.")
            return json.loads(m.group(0))
        except anthropic.RateLimitError:
            wait = 30 * (attempt + 1)
            if attempt < 3:
                st.warning(f"Rate limit — waiting {wait}s before retry {attempt+2}/4…")
                time.sleep(wait)
            else:
                raise
        except anthropic.APIStatusError as e:
            if e.status_code == 529:
                wait = 15 * (attempt + 1)
                if attempt < 3:
                    st.warning(f"Claude is busy — waiting {wait}s before retry {attempt+2}/4…")
                    time.sleep(wait)
                else:
                    st.error("Claude is currently overloaded. Please wait a minute and try again.")
                    raise
            else:
                raise


# ── Excel builder ─────────────────────────────────────────────────────────────
def style_sheet(ws, headers, col_widths):
    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    hdr_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell_font = Font(name="Arial", size=10)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="D0D0D0")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1 blank
    ws.append([""] * len(headers))
    ws.row_dimensions[1].height = 6

    # Row 2 headers
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=ci)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = border
    ws.row_dimensions[2].height = 30

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    return hdr_fill, hdr_font, cell_font, center, left, border


def build_excel(stock_list: list, etf_list: list) -> bytes:
    wb = openpyxl.Workbook()

    # ── Single Stock sheet ──
    ws1 = wb.active
    ws1.title = "Single Stock"
    cell_font = Font(name="Arial", size=10)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="D0D0D0")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    stock_widths = [6,35,45,32,12,6,8,35,30,18,14,30,45,30,50,45,45,8,4,10,35,28]
    style_sheet(ws1, STOCK_HEADERS, stock_widths)

    for rn, s in enumerate(stock_list, start=3):
        row = [
            s.get("run","N"), s.get("companyName",""), s.get("fullCompanyName",""),
            s.get("exchangeName",""), s.get("drTicker",""), "•", s.get("ratio",""),
            s.get("address1",""), s.get("address2",""), s.get("tel",""), s.get("fax",""),
            s.get("companyWebsite",""), s.get("marketNameWebsite",""),
            s.get("marketWebsiteShort",""), s.get("ulMarketWebsite",""),
            s.get("ulIrWebsite",""), s.get("ulIrNews",""),
            s.get("atoFee",0.4), "", s.get("period",""),
            s.get("companyNameTitle",""), s.get("latestPrice",""),
        ]
        ws1.append(row)
        for ci in range(1, len(STOCK_HEADERS)+1):
            c = ws1.cell(row=rn, column=ci)
            c.font = cell_font; c.border = border
            c.alignment = center if ci in (1,5,6,7) else left
        ws1.row_dimensions[rn].height = 20

    # ── ETF sheet ──
    ws2 = wb.create_sheet("ETF")
    etf_widths = [6,35,32,35,14,6,8,35,30,18,14,30,45,30,18,35,40,45,12,18,20,35,8,4,10,35,28]
    style_sheet(ws2, ETF_HEADERS, etf_widths)

    for rn, e in enumerate(etf_list, start=3):
        row = [
            e.get("run","N"), e.get("etfName",""), e.get("exchangeName",""),
            e.get("issuer",""), e.get("drTicker",""), "•", e.get("ratio",""),
            e.get("address1",""), e.get("address2",""), e.get("tel",""), e.get("fax",""),
            e.get("etfWebsite",""), e.get("marketNameWebsite",""),
            e.get("marketWebsiteShort",""), e.get("etfTradingTicker",""),
            e.get("investmentManager",""), e.get("distributor",""),
            e.get("underlyingIndex",""), e.get("baseCurrency",""),
            e.get("inceptionDate",""), e.get("distributionFrequency",""),
            e.get("distributorWebsite",""),
            e.get("atoFee",0.4), "", e.get("period",""),
            e.get("etfNameTitle",""), e.get("latestPrice",""),
        ]
        ws2.append(row)
        for ci in range(1, len(ETF_HEADERS)+1):
            c = ws2.cell(row=rn, column=ci)
            c.font = cell_font; c.border = border
            c.alignment = center if ci in (1,5,6,7) else left
        ws2.row_dimensions[rn].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Session state ─────────────────────────────────────────────────────────────
if "stock_list" not in st.session_state:
    st.session_state.stock_list = []
if "etf_list" not in st.session_state:
    st.session_state.etf_list = []

# ── Styles ────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; }
    .badge { display:inline-block; padding:2px 8px; border-radius:4px;
             font-size:11px; font-weight:700; letter-spacing:.05em; margin-bottom:4px; }
    .badge-col   { background:#1f3864; color:#fff; }
    .badge-price { background:#1a3a1a; color:#6abf6a; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("## 📋 DR Filing Autopilot")
st.caption("Claude AI + Web Search → Review & Edit → Build List → Download Excel")
st.divider()

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    _secret_key = ""
    try:
        _secret_key = st.secrets.get("ANTHROPIC_API_KEY", "")
    except Exception:
        pass
    api_key = _secret_key or os.environ.get("ANTHROPIC_API_KEY", "")

    st.markdown("### 🔑 Anthropic API Key")
    if api_key:
        st.success("API key loaded ✓", icon="🔒")
    else:
        api_key = st.text_input("API Key", type="password", placeholder="sk-ant-...", label_visibility="collapsed")
        if not api_key:
            st.caption("Add to Streamlit secrets: `ANTHROPIC_API_KEY = 'sk-ant-...'`")

    st.divider()
    total = len(st.session_state.stock_list) + len(st.session_state.etf_list)
    st.markdown(f"### 📂 List ({len(st.session_state.stock_list)} stocks · {len(st.session_state.etf_list)} ETFs)")

    if st.session_state.stock_list:
        st.markdown("**Stocks:**")
        for i, s in enumerate(st.session_state.stock_list):
            c1, c2 = st.columns([5,1])
            c1.markdown(f"**{i+1}.** {s.get('companyName','—')}")
            if c2.button("✕", key=f"sdel_{i}"):
                st.session_state.stock_list.pop(i)
                st.rerun()

    if st.session_state.etf_list:
        st.markdown("**ETFs:**")
        for i, e in enumerate(st.session_state.etf_list):
            c1, c2 = st.columns([5,1])
            c1.markdown(f"**{i+1}.** {e.get('etfName','—')}")
            if c2.button("✕", key=f"edel_{i}"):
                st.session_state.etf_list.pop(i)
                st.rerun()

    if not st.session_state.stock_list and not st.session_state.etf_list:
        st.caption("No items added yet.")

    st.divider()
    if total > 0:
        st.download_button(
            "⬇️ Download Excel",
            data=build_excel(st.session_state.stock_list, st.session_state.etf_list),
            file_name=f"DR_Filing_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary",
        )
        if st.button("🗑️ Clear All", use_container_width=True):
            st.session_state.stock_list = []
            st.session_state.etf_list = []
            st.rerun()

# ── Ticker guide ──────────────────────────────────────────────────────────────
with st.expander("📖 How to enter ticker symbols", expanded=False):
    st.markdown("""
| Exchange | Format | Examples |
|---|---|---|
| 🇺🇸 NYSE / NASDAQ (USA) | `TICKER` | `AAPL` · `TSLA` · `COIN` · `CAT` |
| 🇯🇵 Tokyo Stock Exchange | `XXXXXX.T` | `7203.T` · `6857.T` · `6146.T` |
| 🇭🇰 Hong Kong (HKEX) | `XXXX.HK` | `9988.HK` · `1772.HK` · `9626.HK` |
| 🇨🇳 Shanghai (SSE) | `XXXXXX.SS` | `688041.SS` |
| 🇨🇳 Shenzhen (SZSE) | `XXXXXX.SZ` | `300476.SZ` |
| 🇫🇷 Euronext Paris | `XX.PA` | `EL.PA` · `SU.PA` |
| 🇬🇧 London (LSE) | `XX.L` | `AZN.L` · `SHEL.L` |
| 🇺🇸 ETFs (NYSE Arca) | `TICKER` | `GLD` · `SPY` · `QQQ` · `TLT` |

> 💡 You can also type the full name e.g. **SPDR Gold Trust** or **Apple Inc** — Claude will find it.
""")

# ── Two tabs: Stock / ETF ─────────────────────────────────────────────────────
tab_stock, tab_etf = st.tabs(["📈 Single Stock", "🗂️ ETF"])

# ════════════════════════════════════════════════════════
# STOCK TAB
# ════════════════════════════════════════════════════════
with tab_stock:
    st.markdown("### 🔍 Look Up a Stock")
    c1, c2 = st.columns([5,1])
    stock_query   = c1.text_input("Stock ticker", placeholder="e.g. AAPL · TSLA · 7203.T · 9988.HK", label_visibility="collapsed", key="sq")
    stock_retrieve = c2.button("Retrieve", type="primary", use_container_width=True, key="sr")

    if stock_retrieve:
        if not api_key:
            st.error("Please add your Anthropic API key in the sidebar.")
        elif not stock_query.strip():
            st.warning("Enter a ticker or company name.")
        else:
            with st.spinner(f"Searching for **{stock_query}**…"):
                try:
                    data = fetch_data(stock_query.strip(), api_key, STOCK_SYSTEM_PROMPT)
                    data.setdefault("run", "N")
                    st.session_state["stock_pending"] = data
                except Exception as e:
                    st.error(f"Error: {e}")

    if "stock_pending" in st.session_state:
        data = st.session_state["stock_pending"]
        st.divider()
        st.markdown("### ✏️ Review & Edit")
        st.info(f"**{data.get('companyName','')}** · {data.get('exchangeNameEn','')} · 💰 {data.get('latestPrice','—')}")

        st.markdown("**Name columns:**")
        n1, n2, n3 = st.columns(3)
        with n1:
            st.markdown('<span class="badge badge-col">COL 1 — ALL CAPS</span>', unsafe_allow_html=True)
            data["companyName"]      = st.text_area("c1", value=data.get("companyName",""),      height=72, label_visibility="collapsed", key="se1")
        with n2:
            st.markdown('<span class="badge badge-col">COL 2 — CAPS + ("Nickname")</span>', unsafe_allow_html=True)
            data["fullCompanyName"]  = st.text_area("c2", value=data.get("fullCompanyName",""),  height=72, label_visibility="collapsed", key="se2")
        with n3:
            st.markdown('<span class="badge badge-col">COL 3 — Title Case</span>', unsafe_allow_html=True)
            data["companyNameTitle"] = st.text_area("c3", value=data.get("companyNameTitle",""), height=72, label_visibility="collapsed", key="se3")

        st.markdown('<span class="badge badge-price">PRICE</span>', unsafe_allow_html=True)
        data["latestPrice"] = st.text_input("Latest Price", value=data.get("latestPrice",""), key="sep")

        st.divider()
        st.markdown("**Filing fields:**")
        ca, cb = st.columns(2)
        with ca:
            data["exchangeName"]       = st.text_input("Exchange name (Thai)",   value=data.get("exchangeName",""),       key="sf1")
            data["drTicker"]           = st.text_input("DR Ticker",              value=data.get("drTicker",""),           key="sf2")
            data["ratio"]              = st.text_input("Ratio",                  value=str(data.get("ratio","")),         key="sf3")
            data["address1"]           = st.text_input("Address 1",              value=data.get("address1",""),           key="sf4")
            data["address2"]           = st.text_input("Address 2",              value=data.get("address2",""),           key="sf5")
            data["tel"]                = st.text_input("Tel",                    value=data.get("tel",""),                key="sf6")
            data["fax"]                = st.text_input("Fax",                    value=data.get("fax",""),                key="sf7")
        with cb:
            data["companyWebsite"]     = st.text_input("Company website",        value=data.get("companyWebsite",""),     key="sf8")
            data["marketNameWebsite"]  = st.text_area("Market name website",     value=data.get("marketNameWebsite",""),  height=72, key="sf9")
            data["marketWebsiteShort"] = st.text_input("Market website short",   value=data.get("marketWebsiteShort",""), key="sf10")
            data["ulMarketWebsite"]    = st.text_input("UL Market website",      value=data.get("ulMarketWebsite",""),    key="sf11")
            data["ulIrWebsite"]        = st.text_input("UL IR website",          value=data.get("ulIrWebsite",""),        key="sf12")
            data["ulIrNews"]           = st.text_input("UL IR News",             value=data.get("ulIrNews",""),           key="sf13")

        cf, cp, cr = st.columns(3)
        data["atoFee"] = cf.text_input("ATO fee", value=str(data.get("atoFee",0.4)), key="sf14")
        data["period"] = cp.text_input("Period",  value=data.get("period",""), placeholder="e.g. Q326", key="sf15")
        data["run"]    = cr.selectbox("Run", ["N","Y"], index=0 if data.get("run","N")=="N" else 1, key="sf16")

        st.divider()
        ba, bb = st.columns([2,1])
        if ba.button("✅ Add to Stock List", type="primary", use_container_width=True, key="sadd"):
            st.session_state.stock_list.append(dict(data))
            del st.session_state["stock_pending"]
            st.rerun()
        if bb.button("✕ Discard", use_container_width=True, key="sdisc"):
            del st.session_state["stock_pending"]
            st.rerun()

    # Stock list preview
    if st.session_state.stock_list:
        st.divider()
        st.markdown(f"### 📋 Stock List — {len(st.session_state.stock_list)} stock(s)")
        rows = []
        for i, s in enumerate(st.session_state.stock_list, 1):
            rows.append({
                "#": i, "Run": s.get("run","N"),
                "Company name": s.get("companyName",""),
                "Full company name": s.get("fullCompanyName",""),
                "Title Case": s.get("companyNameTitle",""),
                "Latest Price": s.get("latestPrice",""),
                "Exchange": s.get("exchangeNameEn",""),
                "DR Ticker": s.get("drTicker",""),
                "Ratio": s.get("ratio",""),
                "Address 1": s.get("address1",""),
                "Address 2": s.get("address2",""),
                "Tel": s.get("tel",""),
                "Company website": s.get("companyWebsite",""),
                "Market name website": s.get("marketNameWebsite",""),
                "Market website short": s.get("marketWebsiteShort",""),
                "UL Market website": s.get("ulMarketWebsite",""),
                "UL IR website": s.get("ulIrWebsite",""),
                "UL IR News": s.get("ulIrNews",""),
                "ATO fee": s.get("atoFee",""),
                "Period": s.get("period",""),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

# ════════════════════════════════════════════════════════
# ETF TAB
# ════════════════════════════════════════════════════════
with tab_etf:
    st.markdown("### 🔍 Look Up an ETF")
    c1, c2 = st.columns([5,1])
    etf_query    = c1.text_input("ETF ticker or name", placeholder="e.g. GLD · SPY · QQQ · SPDR Gold Trust", label_visibility="collapsed", key="eq")
    etf_retrieve = c2.button("Retrieve", type="primary", use_container_width=True, key="er")

    if etf_retrieve:
        if not api_key:
            st.error("Please add your Anthropic API key in the sidebar.")
        elif not etf_query.strip():
            st.warning("Enter an ETF ticker or name.")
        else:
            with st.spinner(f"Searching for **{etf_query}**…"):
                try:
                    data = fetch_data(etf_query.strip(), api_key, ETF_SYSTEM_PROMPT)
                    data.setdefault("run", "N")
                    st.session_state["etf_pending"] = data
                except Exception as e:
                    st.error(f"Error: {e}")

    if "etf_pending" in st.session_state:
        data = st.session_state["etf_pending"]
        st.divider()
        st.markdown("### ✏️ Review & Edit")
        st.info(f"**{data.get('etfName','')}** · {data.get('exchangeNameEn','')} · 💰 {data.get('latestPrice','—')}")

        st.markdown("**ETF Name:**")
        n1, n2 = st.columns(2)
        with n1:
            st.markdown('<span class="badge badge-col">ETF Name — ALL CAPS</span>', unsafe_allow_html=True)
            data["etfName"]      = st.text_area("en1", value=data.get("etfName",""),      height=72, label_visibility="collapsed", key="ee1")
        with n2:
            st.markdown('<span class="badge badge-col">ETF Name — Title Case</span>', unsafe_allow_html=True)
            data["etfNameTitle"] = st.text_area("en2", value=data.get("etfNameTitle",""), height=72, label_visibility="collapsed", key="ee2")

        st.markdown('<span class="badge badge-price">LATEST NAV / PRICE</span>', unsafe_allow_html=True)
        data["latestPrice"] = st.text_input("Latest NAV/Price", value=data.get("latestPrice",""), key="eep")

        st.divider()
        st.markdown("**Filing fields:**")
        ca, cb = st.columns(2)
        with ca:
            data["exchangeName"]        = st.text_input("Exchange name (Thai)",      value=data.get("exchangeName",""),        key="ef1")
            data["issuer"]              = st.text_input("Issuer",                    value=data.get("issuer",""),              key="ef2")
            data["drTicker"]            = st.text_input("DR Ticker",                 value=data.get("drTicker",""),            key="ef3")
            data["ratio"]               = st.text_input("Ratio",                     value=str(data.get("ratio","")),          key="ef4")
            data["etfTradingTicker"]    = st.text_input("ETF Trading Ticker",        value=data.get("etfTradingTicker",""),    key="ef5")
            data["address1"]            = st.text_input("Address 1",                 value=data.get("address1",""),            key="ef6")
            data["address2"]            = st.text_input("Address 2",                 value=data.get("address2",""),            key="ef7")
            data["tel"]                 = st.text_input("Tel",                       value=data.get("tel",""),                 key="ef8")
            data["fax"]                 = st.text_input("Fax",                       value=data.get("fax",""),                 key="ef9")
            data["etfWebsite"]          = st.text_input("ETF Website",               value=data.get("etfWebsite",""),          key="ef10")
        with cb:
            data["marketNameWebsite"]   = st.text_area("Market name website",        value=data.get("marketNameWebsite",""),   height=72, key="ef11")
            data["marketWebsiteShort"]  = st.text_input("Market website short",      value=data.get("marketWebsiteShort",""),  key="ef12")
            data["investmentManager"]   = st.text_input("Investment Manager",        value=data.get("investmentManager",""),   key="ef13")
            data["distributor"]         = st.text_input("Distributor",               value=data.get("distributor",""),         key="ef14")
            data["underlyingIndex"]     = st.text_area("Underlying reference index", value=data.get("underlyingIndex",""),     height=72, key="ef15")
            data["baseCurrency"]        = st.text_input("Base Currency",             value=data.get("baseCurrency",""),        key="ef16")
            data["inceptionDate"]       = st.text_input("Inception Date",            value=data.get("inceptionDate",""),       key="ef17")
            data["distributionFrequency"]= st.text_input("Distribution Frequency",  value=data.get("distributionFrequency",""),key="ef18")
            data["distributorWebsite"]  = st.text_input("Distributor Website",       value=data.get("distributorWebsite",""),  key="ef19")

        cf, cp, cr = st.columns(3)
        data["atoFee"] = cf.text_input("ATO fee", value=str(data.get("atoFee",0.4)), key="ef20")
        data["period"] = cp.text_input("Period",  value=data.get("period",""), placeholder="e.g. Q326", key="ef21")
        data["run"]    = cr.selectbox("Run", ["N","Y"], index=0 if data.get("run","N")=="N" else 1, key="ef22")

        st.divider()
        ba, bb = st.columns([2,1])
        if ba.button("✅ Add to ETF List", type="primary", use_container_width=True, key="eadd"):
            st.session_state.etf_list.append(dict(data))
            del st.session_state["etf_pending"]
            st.rerun()
        if bb.button("✕ Discard", use_container_width=True, key="edisc"):
            del st.session_state["etf_pending"]
            st.rerun()

    # ETF list preview
    if st.session_state.etf_list:
        st.divider()
        st.markdown(f"### 📋 ETF List — {len(st.session_state.etf_list)} ETF(s)")
        rows = []
        for i, e in enumerate(st.session_state.etf_list, 1):
            rows.append({
                "#": i, "Run": e.get("run","N"),
                "ETF Name": e.get("etfName",""),
                "ETF Name (Title Case)": e.get("etfNameTitle",""),
                "Latest Price": e.get("latestPrice",""),
                "Exchange": e.get("exchangeNameEn",""),
                "Issuer": e.get("issuer",""),
                "DR Ticker": e.get("drTicker",""),
                "Ratio": e.get("ratio",""),
                "Trading Ticker": e.get("etfTradingTicker",""),
                "Investment Manager": e.get("investmentManager",""),
                "Distributor": e.get("distributor",""),
                "Underlying Index": e.get("underlyingIndex",""),
                "Base Currency": e.get("baseCurrency",""),
                "Inception Date": e.get("inceptionDate",""),
                "Distribution Freq": e.get("distributionFrequency",""),
                "ATO fee": e.get("atoFee",""),
                "Period": e.get("period",""),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

st.divider()
st.caption("👈 Click **Download Excel** in the sidebar — exports both Single Stock and ETF sheets.")
